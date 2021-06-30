
import configparser
from os import truncate
from socket import *
import time
import threading 
import os

import xlrd
import math
import numpy as np 
import xlsxwriter
from  openpyxl import  Workbook 

from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.pyplot import MultipleLocator
from datetime import datetime
from xlrd import xldate_as_tuple

import pymysql


#开发者日志：
#2021/4/20 白马腾 完成excel读入与显示化调试部分
#2021/4/23 白马腾 完成自动识别长度系统，允许程序检验支架数据报错， 
#2021/4/27 白马腾 完成一阶假设运算函数
#2021/5/2  白马腾 完成图形化显示 对应点校准 支架直径变化量分析（径向） 支架旋转角度计算（周向） 欠缺：支架绝对位移计算 限制型支架交叠分析
#2021/5/3  白马腾 完成数据绝对位置计算 所有空间形态学特征计算完毕 可以开始进行分析 完成控制器构
#2021/5/9  白马腾 使用属性长度方法更换calculate运算标识符 进行第一次算法重构 解决重叠标识与方向问题   欠缺：支架首末时期的整体性分析
#2021/5/11 白马腾 重构基本方法，允许进行自定义时间访问 欠缺支架环节长度 支架环节夹角 支架整体夹角和趋势 数据分析
#2021/5/13 白马腾 进行支架环环节夹角，支架环节长度计算 支架运动的可视化描述
#2021/5/16 白马腾 进行算法重构 建立新的数据库模型 拓展接口的适应性 拓展支架分析的完备性数据集
#2021/5/24 白马腾 重构数据录入方法 使用huan数组 新加入8点型环
#2021/5/28 白马腾 完成血管信息录入
#2021/6/13 白马腾 V1版本，仅对基本功能进行描述  提供数据分析与数据处理 使用LAN全连接的方式进行信息分布式聚集 传递的本质是广播
#2021/6/15 白马腾 V2版本，引入洪泛结构 可以动态配置传输路径 
#2021/6/18 白马腾 DB版本，使用mysql数据库进行数据管理
#2021/6/19 白马腾 DB2版本，添加数据库模块中 角色 视图 权限管理 函数 数据可视化 操作

#【网络数据传输类定义】

class inf():
    def __init__(self,ip,sock,id):
        self.ip=ip
        self.sock=sock
        self.id=id


#【全局信息定义】

#shape反映了在支架的每个点处取数据点的个数
stent_shape0=[8,5,5,5,5,5,5,5,5,5,8,8]            #限制型近端支架[11节长支架] 
stent_shape1=[4,4,4,4,4,4,4,4,4]                  #限制型远端支架[8节短支架]
stent_shape2=[8,5,5,5,5,5,5,5,8,8]                #限制型近端支架[9节长支架]
#适配老数据
stent_shape3=[4,5,5,5,5,5,5,5,5,5,4,4]            #限制型近端支架[11节长支架] 
stent_shape4=[4,5,5,5,5,5,5,5,4,4]                #限制型近端支架[9节长支架]
#支持快速访问数据
# stent_type=1  近端支架     stent_type=2  远端支架 
# stent_shape=0 11环近端支架 stent_shape=1 8环远端支架 stent_shape=2 9环近端支架 
ALL_stent_type=["","近端支架","远端支架"]
ALL_stent_shape=["11环近端支架","8环远端支架","9环近端支架"]
#支架形状总列表
stent_shape_list=[]                               
stent_shape_list.append(stent_shape0)
stent_shape_list.append(stent_shape1)
stent_shape_list.append(stent_shape2)
stent_shape_list.append(stent_shape3)
stent_shape_list.append(stent_shape4)
#本地文件夹路径设置    [注意route_0是 \ 形式]  [使用replace进行代换]
route_0=os.path.dirname(os.path.realpath(__file__))
route=route_0.replace("\\","\\\\")
#ini文件读取
inipath = route+"\\\\org.ini"           # org.ini的路径               
conf = configparser.ConfigParser()      # 创建管理对象
conf.read(inipath, encoding="utf-8")    # 读取ini文件
items = conf.items('ip_socket_set')     # 读入对应section
#写入表格数据标记量
write_number=0
#本地数据表patient_list=[]
patient_list=[]
#自身信息：
inf_self=inf(items[0][1],int(items[1][1]),items[2][1])
#组建LAN动态IP表                         # 用于储存本地inf数据 与被链接inf数据
inf_list=[]
inf_list.append(inf_self)               # 加入自身信息
#组建临时LAN接收IP表                     用于储存负反馈广播所接收到的inf数据
inf_temp_list=[]
#组建洪泛信息列表                        # 用于维护洪泛网络动态链接inf数据
inf_flooding_list=[]
items = conf.items('inf_flooding_list_set')     # 读入对应section

for i in range(0,int(len(items)/2)):            # 读取原始配置文件加入洪泛列表
    ip_temp=items[2*i][1]
    socket_temp=int(items[2*i+1][1])
    id_temp="unknown"
    inf_temp=inf(ip_temp,socket_temp,id_temp)
    inf_flooding_list.append(inf_temp)

#用户权限变量
jurisdiction=[]



#【支架统计计算公用函数定义】

#给出以 1992/8/3 形式的日期数据 返回python日期数据值
def HowLong(stime,etime):
    start = datetime.strptime(stime, '%Y/%m/%d')
    end = datetime.strptime(etime, '%Y/%m/%d')
    time=(end - start).total_seconds()/(60*60*24)                  #秒数除以60*60*24算出花费的天数
    return time

# 输入四点坐标A,B,C,D 进行矢量点积[矢量BA*矢量DC] 返回所得值 
def dot(pointA,pointB,pointC,pointD):
    r1=pointA.x-pointB.x
    r2=pointA.y-pointB.y
    r3=pointA.z-pointB.z
    r4=pointC.x-pointD.x
    r5=pointC.y-pointD.y
    r6=pointC.z-pointD.z
    return (r1*r4+r2*r5+r3*r6)

# 输入四点坐标A,B,C,D 进行矢量点积[矢量BA*矢量DC] 返回夹角cos值 
def dot_cos(pointA,pointB,pointC,pointD):
    r1=pointA.x-pointB.x
    r2=pointA.y-pointB.y
    r3=pointA.z-pointB.z
    r4=pointC.x-pointD.x
    r5=pointC.y-pointD.y
    r6=pointC.z-pointD.z
    return (r1*r4+r2*r5+r3*r6)/((r1*r1+r2*r2+r3*r3)**0.5*(r4*r4+r5*r5+r6*r6)**0.5)

# 输入四点坐标A,B,C,D 进行矢量点积[矢量BA*矢量DC] 返回夹角角度值大小
def angle(pointA,pointB,pointC,pointD):
    r1=pointA.x-pointB.x
    r2=pointA.y-pointB.y
    r3=pointA.z-pointB.z
    r4=pointC.x-pointD.x
    r5=pointC.y-pointD.y
    r6=pointC.z-pointD.z
    angle_cos=(r1*r4+r2*r5+r3*r6)/((r1*r1+r2*r2+r3*r3)**0.5*(r4*r4+r5*r5+r6*r6)**0.5)
    angle_temp=math.acos(angle_cos)*180/math.pi
    return angle_temp

#输入两点坐标返回距离
def distance(pointA,pointB):
    return ((pointA.x-pointB.x)**2+(pointA.y-pointB.y)**2+(pointA.z-pointB.z)**2)**0.5

# 输入四点坐标A,B,C,D 进行矢量叉积[矢量BA 叉乘 矢量DC] 返回原点point值 [原点矢量]
def cross(pointA,pointB,pointC,pointD):
    r1=pointA.x-pointB.x
    r2=pointA.y-pointB.y
    r3=pointA.z-pointB.z
    r4=pointC.x-pointD.x
    r5=pointC.y-pointD.y
    r6=pointC.z-pointD.z
    return point(r2*r6-r3*r5,r3*r4-r1*r6,r1*r5-r2*r4)

# 减函数 输入A,B两点 输出BA向量点
def jian(pointA,pointB):
    return point(pointA.x-pointB.x,pointA.y-pointB.y,pointA.z-pointB.z)

# 加函数 输入A,B两点 输出加和向量点
def jia(pointA,pointB):
    return point(pointA.x+pointB.x,pointA.y+pointB.y,pointA.z+pointB.z)



# 【支架统计计算类定义】

#定义点类
class point():
    def __init__(self,x,y,z):
        #变量初始化
        self.x=x
        self.y=y
        self.z=z
        return 

#定义环类【注:环类坐标使用直接坐标,在period间运算时引入骨骼坐标】
class huan():
    def __init__(self,number_points,point_list_temp):
        #初始化环上点
        self.number_points=number_points                    #记录环类型 1点型 4点型 5点型 8点型

        #不同条件下选用不同方式进行数据访问                    #记录各点数据
        self.point_list=point_list_temp                     
        self.point_mid=point(0,0,0)
        self.flag=" "                                       #记录该环是否重叠 非重叠部分：“ ”   重叠部分：“*”
        self.r_list=[]                                      #记录模拟环半径列表

        #将数组中坐标求平均值计算point_mid
        for i in range(0,number_points):
            self.point_mid.x=self.point_mid.x+point_list_temp[i].x/number_points
            self.point_mid.y=self.point_mid.y+point_list_temp[i].y/number_points
            self.point_mid.z=self.point_mid.z+point_list_temp[i].z/number_points

        #使用中心点估算环截面半径分布情况
        for i in range(0,number_points):
            r_temp=distance(self.point_mid,self.point_list[i])
            self.r_list.append(r_temp)

        return

#定义支架类
class stent():
    def __init__(self,stent_type,stent_shape):

        # [stent_type=0 原始血管 1.0版本已移除该选项] stent_type=1 近端支架 stent_type=2 远端支架 
        # stent_shape=0 11环近端支架 stent_shape=1 8环远端支架 stent_shape=2 9环近端支架 

        self.stent_type=stent_type              #stent_type定义支架类型
        self.stent_shape=stent_shape            #stent_type定义支架环分布类型
        self.huan_list=[]                       #huan_list用于存放各环
        self.zhou_angle_list=[]                 #存放支架弯角
        self.zhou_angle_all=0                   #存放轴向弯角和
        return

    def add_huan(self,huan001):                 #增添huan函数默认增添顺序由上至下
        self.huan_list.append(huan001)              
        return      

#定义血管观察点类
class see_point():
    #Px,Py,Pz:    coordinates of point P on the centerline
    #Tx,Ty,Tz:    coordinates of the direction of the Tangent vector to point P
    #Nx,Ny,Nz:    coordinates of the direction of the Normal vector to point P
    #BNx,BNy,BNz: coordinates of the direction of the Binormal vector to point P
    #Dfit:        diameter of the best fit circle in point P
    #Dmin:        diameter of the inscribing circle in point P
    #Dmax:        diameter of the subscribing circle in point P
    #C:           curvature in point P
    #Dh:          hydraulic diameter in point P
    #Xh:          hydraulic ratio in point P
    #Scf:         distance of circumference in point P
    #Area:        sectional area in point P
    #E:           ellipticity in point P
    def __init__(self,x,y,z,Tx,Ty,Tz,Nx,Ny,Nz,BNx,BNy,BNz,Dfit,Dmin,Dmax,C,Dh,Xh,Scf,Area,E) :
        self.x=x
        self.y=y
        self.z=z
        self.Tx=Tx
        self.Ty=Ty
        self.Tz=Tz
        self.Nx=Nx
        self.Ny=Ny
        self.Nz=Nz
        self.BNx=BNx
        self.BNy=BNy
        self.BNz=BNz
        self.Dfit=Dfit
        self.Dmin=Dmin
        self.Dmax=Dmax
        self.C=C
        self.Dh=Dh
        self.Xh=Xh
        self.Scf=Scf
        self.Area=Area
        self.E=E
        self.information_list=[]
        self.information_list.append(x)
        self.information_list.append(y)
        self.information_list.append(z)
        self.information_list.append(Tx)
        self.information_list.append(Ty)
        self.information_list.append(Tz)
        self.information_list.append(Nx)
        self.information_list.append(Ny)
        self.information_list.append(Nz)
        self.information_list.append(BNx)
        self.information_list.append(BNy)
        self.information_list.append(BNz)
        self.information_list.append(Dfit)
        self.information_list.append(Dmin)
        self.information_list.append(Dmax)
        self.information_list.append(C)
        self.information_list.append(Dh)
        self.information_list.append(Xh)
        self.information_list.append(Scf)
        self.information_list.append(Area)
        self.information_list.append(E)

    def show(self):         #提供展示校验方法
        print(self.x,
            self.y,
            self.z,
            self.Tx,
            self.Ty,
            self.Tz,
            self.Nx,
            self.Ny,
            self.Nz,
            self.BNx,
            self.BNy,
            self.BNz,
            self.Dfit,
            self.Dmin,
            self.Dmax,
            self.C,
            self.Dh,
            self.Xh,
            self.Scf,
            self.Area,
            self.E)

#定义时期类
class period():
    def __init__(self,data,top12_x,top12_y,top12_z):
        self.data=data                             #定义支架时期
        self.top12_x=top12_x                       #定义骨骼位置
        self.top12_y=top12_y                       #定义骨骼位置
        self.top12_z=top12_z                       #定义骨骼位置
        self.top12=point(top12_x,top12_y,top12_z)  #定义骨骼位置点
        self.stent_list=[]                         #增加支架列表
        #储存观察点受力反解信息
        self.Fx=[]
        self.Fy=[]
        self.Fz=[]
        self.F=[]
        #储存观察点约束信息
        self.kx=[]
        self.ky=[]
        self.kz=[]
        #储存血管信息
        self.see_point_list=[]
        return

    def add_stent(self,stent001):               #默认由上至下添加支架
        self.stent_list.append(stent001)
        return

    def show_3D(self):
        fig = plt.figure()
        ax1 = plt.axes(projection='3d')

        plt.title(self.data,fontsize=24)

        for i in range(0,len(self.stent_list)):
            list_xtemp=[]
            list_ytemp=[]
            list_ztemp=[]
            for j in range(0,len(self.stent_list[i].huan_list)):
                list_xtemp.append(self.stent_list[i].huan_list[j].point_mid.x)
                list_ytemp.append(self.stent_list[i].huan_list[j].point_mid.y)
                list_ztemp.append(self.stent_list[i].huan_list[j].point_mid.z)

                list_x_huantemp=[]
                list_y_huantemp=[]
                list_z_huantemp=[]

                for k in range(0,self.stent_list[i].huan_list[j].number_points):
                    list_x_huantemp.append(self.stent_list[i].huan_list[j].point_list[k].x)
                    list_y_huantemp.append(self.stent_list[i].huan_list[j].point_list[k].y)
                    list_z_huantemp.append(self.stent_list[i].huan_list[j].point_list[k].z)

                list_x_huantemp.append(self.stent_list[i].huan_list[j].point_list[0].x)
                list_y_huantemp.append(self.stent_list[i].huan_list[j].point_list[0].y)
                list_z_huantemp.append(self.stent_list[i].huan_list[j].point_list[0].z)

                ax1.scatter3D(list_x_huantemp,list_y_huantemp,list_z_huantemp, cmap='Blues')   #绘制散点图
                ax1.plot3D(list_x_huantemp,list_y_huantemp,list_z_huantemp,'gray')             #绘制空间曲线
     

            ax1.scatter3D(list_xtemp,list_ytemp,list_ztemp, cmap='Blues')   #绘制散点图
            ax1.plot3D(list_xtemp,list_ytemp,list_ztemp,'gray')             #绘制空间曲线     
        plt.show()  
        return 
        
#定义病例类
class patient():
    def __init__(self,sex,name,ID):
        self.sex=sex
        self.name=name
        self.ID=ID
        self.period_list=[]                     #储存病人各时期数据
        self.calculation=False                  #默认病人数据处于不可运算状态【缺失病人原始血管数据】
        self.period_length=0                    #时期数组遍历长度
        self.xianzhixing=False                  #默认病人是非限制性支架病例 搜索到限制型支架时 进行更换
        return
    
    def add_period(self,period001):             #默认病例时间由早至晚
        self.period_list.append(period001)
        if len(period001.stent_list)!=0:
            self.period_length=self.period_length+1
            return 
        else:
            return

    def show(self):
        print("|~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|")
        print("|name : ",self.name)
        print("|sex  : ",self.sex)
        print("|ID   : ",self.ID)
        print("|     |***************************************************************|")
        for i in range(0,len(self.period_list)):
            print("|     |period",i+1)
            print("|     |period data : ",self.period_list[i].data)
            print("|     |Relative positioning point of bone : ")
            print("|     |[",self.period_list[i].top12_x,",",self.period_list[i].top12_y,",",self.period_list[i].top12_z,"]")
            print("|     |     |---------------------------------------------------------|")
            for j in range(0,len(self.period_list[i].stent_list)):
                if self.period_list[i].stent_list[j].stent_shape==0:
                    str_temp="8-5-8型11环支架"
                elif self.period_list[i].stent_list[j].stent_shape==1:
                    str_temp="8-8型8环支架"
                elif self.period_list[i].stent_list[j].stent_shape==2:
                    str_temp="8-5-8型9环支架"
                elif self.period_list[i].stent_list[j].stent_shape==3:
                    str_temp="旧版8-5-8型11环支架"
                elif self.period_list[i].stent_list[j].stent_shape==4:
                    str_temp="旧版8-5-8型9环支架"
                else:
                    str_temp="未定义的支架类型"
                print("|     |     |stent_type:",str_temp)
                for k in range(0,len(self.period_list[i].stent_list[j].huan_list)): 
                    print("|     |     |point_mid [",round(self.period_list[i].stent_list[j].huan_list[k].point_mid.x,2),\
                          ",",round(self.period_list[i].stent_list[j].huan_list[k].point_mid.y,2),\
                          ",",round(self.period_list[i].stent_list[j].huan_list[k].point_mid.z,2),"]")
                print("|     |     |---------------------------------------------------------|")
            print("|     |***************************************************************|")
        print("|~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|")
        return

    def show_F(self):
        if self.calculation==True:
            print("|~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|")
            print("|name : ",self.name)
            print("|sex  : ",self.sex)
            print("|ID   : ",self.ID)
            for i in range(0,len(self.period_list[-1].stent_list[0].huan_list)):
                print("第 ",i+1," 个观察点 F : ",round(self.period_list[0].F[i],2), "Fx: ",round(self.period_list[0].Fx[i],2)," Fy: ",round(self.period_list[0].Fy[i],2)," Fz: ",round(self.period_list[0].Fz[i],2))
        return

    def show_k(self):
        if self.calculation==True:
            print("|~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|")
            print("|name : ",self.name)
            print("|sex  : ",self.sex)
            print("|ID   : ",self.ID)
            for i in range(0,len(self.period_list[-1].stent_list[0].huan_list)):
                print("第 ",i+1," 个观察点 kx: ",round(self.period_list[0].kx[i],2)," ky: ",round(self.period_list[0].ky[i],2)," kz: ",round(self.period_list[0].kz[i],2))
        return

    def write(self,route_temp,file_name):  #在route_temp路径追加写入该病例信息
        workbook = xlsxwriter.Workbook(route+"\\"+file_name)
        worksheet = workbook.add_worksheet('Sheet1')
        number_x=0                #记录x写入位置 从第一列开始读入        
        number_y=1                #记录y写入位置 从第二行开始读入
        worksheet.write(number_y,number_x,self.name)    #写入姓名
        worksheet.write(number_y,number_x+1,self.ID)    #写入ID
        worksheet.write(number_y,number_x+2,self.sex)   #写入性别

        for i in range(0,self.period_length):
            number_x=3                                      #x归位
            date_time_temp = datetime.strptime(self.period_list[i].data, '%Y/%m/%d')
            worksheet.write_datetime(number_y+i,number_x,date_time_temp)                    #写入日期
            worksheet.write(number_y+i,number_x+1,self.period_list[i].top12_x)              #写入x
            worksheet.write(number_y+i,number_x+2,self.period_list[i].top12_y)              #写入y
            worksheet.write(number_y+i,number_x+3,self.period_list[i].top12_z)              #写入z
            number_x=number_x+4                             #x移位
            #支架写入部分
            for i1 in range(0,len(self.period_list[i].stent_list)): #遍历所有支架
                worksheet.write(number_y+i,number_x,"#")    #写入#
                worksheet.write(number_y+i,number_x+1,self.period_list[i].stent_list[i1].stent_type)
                worksheet.write(number_y+i,number_x+2,self.period_list[i].stent_list[i1].stent_shape)
                number_x=number_x+3                         #x移位
                for i2 in range(0,len(self.period_list[i].stent_list[i1].huan_list)):
                    huan_temp=self.period_list[i].stent_list[i1].huan_list[i2]
                    for i3 in range(0,len(huan_temp.point_list)):
                        point_temp=huan_temp.point_list[i3]
                        worksheet.write(number_y+i,number_x,point_temp.x) 
                        worksheet.write(number_y+i,number_x+1,point_temp.y)
                        worksheet.write(number_y+i,number_x+2,point_temp.z)
                        number_x=number_x+3                 #x移位
            #血管数据写入部分
            #x,y,z,Tx,Ty,Tz,Nx,Ny,Nz,BNx,BNy,BNz,Dfit,Dmin,Dmax,C,Dh,Xh,Scf,Area,E
            if len(self.period_list[i].see_point_list)!=0:  #若该时期存在有血管数据
                worksheet.write(number_y+i,number_x,"**")
                number_x=number_x+1                 #x移位
                for i1 in range(0,len(self.period_list[i].see_point_list)):
                    see_point_temp=self.period_list[i].see_point_list[i1]
                    for i2 in range(0,len(see_point_temp.information_list)):
                        worksheet.write(number_y+i,number_x+i2,see_point_temp.information_list[i2])
                    number_x=number_x+len(see_point_temp.information_list)                 #x移位
            #末位标识写入
            worksheet.write(number_y+i,number_x,"##")
        number_x=3                                      #x归位
        number_y=number_y+self.period_length            #y移位
        #第零时期写入
        if self.calculation==True:              #若存在第零时期
            date_time_temp = datetime.strptime(self.period_list[i].data, '%Y/%m/%d')
            worksheet.write_datetime(number_y,number_x,date_time_temp)           #写入日期
            worksheet.write(number_y,number_x+1,self.period_list[-1].top12_x)    #写入x
            worksheet.write(number_y,number_x+2,self.period_list[-1].top12_y)    #写入y
            worksheet.write(number_y,number_x+3,self.period_list[-1].top12_z)    #写入z
            number_x=number_x+4                                      #x移位
            worksheet.write(number_y,number_x,"**")                #写入标识符
            number_x=number_x+1                                      #x移位
            for i1 in range(0,len(self.period_list[-1].see_point_list)):
                    see_point_temp=self.period_list[-1].see_point_list[i1]
                    for i2 in range(0,len(see_point_temp.information_list)):
                        worksheet.write(number_y,number_x+i2,see_point_temp.information_list[i2])
                    number_x=number_x+len(see_point_temp.information_list)                 #x移位
            #末位标识写入
            worksheet.write(number_y,number_x,"##")
        workbook.close()
        return 


                            
                    


        for i in range(len(data)-1,flag,-1):
            list_temp=data[i].split()
            for j in range(0,len(list_temp)):
                worksheet.write(0,number,list_temp[j])
                number=number+1
        worksheet.write(0,number,"##")
        workbook.close()

#文件追加算法
def read(route_temp): #将route文件读入本地patient_list中
    #主函数运算
    book01 = xlrd.open_workbook(route_temp)
    sheet=book01.sheets()[0]                                    #默认数据在sheet1中
    nrows=sheet.nrows                                           #得到总行数
    ncols=sheet.ncols                                           #得到总列数

    #读取database
    
    i=1
    while True:                                                 #第一行为属性名
        if nrows<=1:                                            #空表，直接退出
            break
        if sheet.cell(i,1).value!='':
            patient_temp=patient(sheet.cell(i,2).value,sheet.cell(i,0).value,sheet.cell(i,1).value)
            temp_i=1                                            #记录区间跳越值
            if i+temp_i<nrows:
                while sheet.cell(i+temp_i,1).value=='' :
                    temp_i=temp_i+1
                    if i+temp_i==nrows:
                        break

        for j in range(i,i+temp_i):    
            period_temp=period(datetime(*xldate_as_tuple(sheet.cell_value(j,3),0)).strftime('%Y/%m/%d'),             #data赋值
                                        sheet.cell(j,4).value,sheet.cell(j,5).value,sheet.cell(j,6).value)           #top-12赋值
            k=7
            while True:
                if sheet.cell(j,k).value=='##' :                # 检索到##末端，结束
                    break
                elif sheet.cell(j,k).value=='**' :              # 检索到* 进入血管数据读取
                    if len(patient_temp.period_list)==0:        # 如果此时对应的时期列表为空，证明该时期是原始血管时期
                        patient_temp.calculation=True
                    while True:
                        see_point_temp=see_point(sheet.cell(j,k+1).value,sheet.cell(j,k+2).value,sheet.cell(j,k+3).value,
                                                sheet.cell(j,k+4).value,sheet.cell(j,k+5).value,sheet.cell(j,k+6).value,
                                                sheet.cell(j,k+7).value,sheet.cell(j,k+8).value,sheet.cell(j,k+9).value,
                                                sheet.cell(j,k+10).value,sheet.cell(j,k+11).value,sheet.cell(j,k+12).value,
                                                sheet.cell(j,k+13).value,
                                                sheet.cell(j,k+14).value,
                                                sheet.cell(j,k+15).value,
                                                sheet.cell(j,k+16).value,
                                                sheet.cell(j,k+17).value,
                                                sheet.cell(j,k+18).value,
                                                sheet.cell(j,k+19).value,
                                                sheet.cell(j,k+20).value,
                                                sheet.cell(j,k+21).value)
                        k=k+21
                        period_temp.see_point_list.append(see_point_temp)
                        if sheet.cell(j,k+1).value=='##':
                            break
                    period_temp.calculate=True
                    break
                elif sheet.cell(j,k).value=='#' :               # 检索到#
                    stent_temp=stent(sheet.cell(j,k+1).value,sheet.cell(j,k+2).value)   # 构造临时支架
                    if sheet.cell(j,k+1).value==2:
                        patient_temp.xianzhixing=True           # 若检测到是限制型支架，则更改标识
                    k=k+2                                       # 将k移位至数据单元格前
                    if sheet.cell(j,k+1).value=='##':           # 空数据判定，若为##跳过直接进行下一行的读取
                        break
                    for p in range(0,len(stent_shape_list[int(stent_temp.stent_shape)])):
                        points_list_temp=[]
                        for q in range(0,stent_shape_list[int(stent_temp.stent_shape)][p]):
                            points_list_temp.append(point(sheet.cell(j,k+q*3+1).value,sheet.cell(j,k+q*3+2).value,sheet.cell(j,k+q*3+3).value))
                        huan_temp=huan(stent_shape_list[int(stent_temp.stent_shape)][p],points_list_temp)           
                        stent_temp.add_huan(huan_temp)                                                                                          
                        k=k+stent_shape_list[int(stent_temp.stent_shape)][p]*3                                                                  #移位k
                    period_temp.add_stent(stent_temp)           #近端支架录入完成向period_temp追加stent_temp
                    k=k+1                                       #将k移位
                else:
                    print("第",j+1,"行","第",k+1,"列 数据格式出错!!!")
                    print(sheet.cell(j,k).value)
                    break
            patient_temp.add_period(period_temp)                #每行扫描完成后向patient中追加period_temp
        #重复性判断:
        biaoshi=0
        for i0 in range(0,len(patient_list)):
            if patient_list[i0].ID.strip()==patient_temp.ID.strip():            #若检测到相同学生
                biaoshi=1
                print("SAME ID : ",patient_temp.ID)
                break
        if biaoshi==0:
            patient_list.append(patient_temp)           #确认无重复时，记录完成后向patient_list追加patient_temp
        i=i+temp_i                                      #i值跳跃
        if i==nrows:
            break

#环数据整理矫正算法
def arrangement():
    '''
    通过求解COS值选取对应关系，即假设偏转角最大不超过36度
    仅对五点型支架进行判定
    '''
    #病人第一时期对应点关系构建
    for i1 in range(0,len(patient_list)):                                                               #遍历各病人
        for i3 in range(0,len(patient_list[i1].period_list[0].stent_list)):                             #对第一时期进行判定
            for i4 in range(0,len(patient_list[i1].period_list[0].stent_list[i3].huan_list)):           #遍历对应环集
                if patient_list[i1].period_list[0].stent_list[i3].huan_list[i4].number_points==5 and i4>0:      #若为5点型环且非首环
                    if patient_list[i1].period_list[0].stent_list[i3].huan_list[i4-1].number_points==5:         #若前环同样为五点型环
                        new_points_list=[]                     #创建赋值总表
                        point1=patient_list[i1].period_list[0].stent_list[i3].huan_list[i4-1].point_mid
                        point2=patient_list[i1].period_list[0].stent_list[i3].huan_list[i4].point_mid
                        for i5 in range(0,5):
                            new_points_list_temp=[]            #创建新一轮比较列表
                            for i6 in range(0,5):
                                new_points_list_temp.append(dot_cos(patient_list[i1].period_list[0].stent_list[i3].huan_list[i4-1].point_list[i5],
                                                                patient_list[i1].period_list[0].stent_list[i3].huan_list[i4].point_list[i6],
                                                                point1,
                                                                point2))
                            r=new_points_list_temp.index(max(new_points_list_temp))
                            new_points_list.append(patient_list[i1].period_list[0].stent_list[i3].huan_list[i4].point_list[r])
                        #循环5次后new_points_list存储五点对应信息 下面进行数据调换
                        for i5 in range(0,5):
                            patient_list[i1].period_list[0].stent_list[i3].huan_list[i4].point_list[i5]=new_points_list[i5]

    #同病人非第一时期时期 对应点关系构建
    for i1 in range(0,len(patient_list)):                                                                #遍历各病人
        for i2 in range(1,patient_list[i1].period_length):                                           #遍历该病人各运算时期
            for i3 in range(0,len(patient_list[i1].period_list[i2].stent_list)):
                for i4 in range(0,len(patient_list[i1].period_list[i2].stent_list[i3].huan_list)):
                    if patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].number_points==5:   #若为5点型环
                        new_points_list=[]                     #创建赋值总表
                        top12_1=patient_list[i1].period_list[i2-1].stent_list[i3].huan_list[i4].point_mid
                        top12_2=patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_mid
                        for i5 in range(0,5):                  
                            new_points_list_temp=[]            #创建新一轮比较列表
                            for i6 in range(0,5):
                                new_points_list_temp.append(dot_cos(patient_list[i1].period_list[i2-1].stent_list[i3].huan_list[i4].point_list[i5],top12_1,
                                                                patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_list[i6]  ,top12_2))
                            r=new_points_list_temp.index(max(new_points_list_temp))
                            new_points_list.append(patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_list[r])
                        #循环5次后new_points_list存储五点对应信息 下面进行数据调换
                        for i5 in range(0,5):
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_list[i5]=new_points_list[i5]

    #患者重叠部分支架的运算:
    #自下向上检索近端支架 将低于远端支架的最高点的环标* 
    #自上向下检索远端支架 将高于近端支架的最低点的环标* 
    for i1 in range(0,len(patient_list)):                                                                #遍历各病人
        if len(patient_list[i1].period_list[0].stent_list)==2 and patient_list[i1].period_list[0].stent_list[0].stent_type==1 and patient_list[i1].period_list[0].stent_list[1].stent_type==2 :
            #若该患者是标准限制型支架                  
            for i2 in range(0,patient_list[i1].period_length):                                     #遍历该病人各运算时期
                z_least=patient_list[i1].period_list[i2].stent_list[0].huan_list[-1].point_mid.z   #找到近端支架的末端高度
                z_top=patient_list[i1].period_list[i2].stent_list[1].huan_list[0].point_mid.z      #找到远端支架的顶端高度
                for i3 in range(len(patient_list[i1].period_list[i2].stent_list[0].huan_list)-1,-1,-1):  #遍历近端支架
                    if patient_list[i1].period_list[i2].stent_list[0].huan_list[i3].point_mid.z<=z_top:
                        patient_list[i1].period_list[i2].stent_list[0].huan_list[i3].flag="*"
                        continue
                    else:
                        break
                for i3 in range(0,len(patient_list[i1].period_list[i2].stent_list[1].huan_list)):  #遍历支架末端环
                    if patient_list[i1].period_list[i2].stent_list[1].huan_list[i3].point_mid.z>=z_least:
                        patient_list[i1].period_list[i2].stent_list[1].huan_list[i3].flag="*"
                        continue
                    else:
                        break

#数据集统计展示代码
def show_database():
    #录入数据的全集打印
    for i in range(0,len(patient_list)):
        patient_list[i].show()
        print()

#数据3D展示代码
def show_3D():
    for i1 in range(0,len(patient_list)):                                                                   #遍历各病人
        for i2 in range(0,len(patient_list[i1].period_list)):                                               #遍历该病人各时期
            patient_list[i1].period_list[i2].show_3D()

#remake_database用于重新计算各项数据
def remake_database():
    remake_zhou_angle()         #重新计算轴向偏角

#支架轴向偏角计算 【同一支架 三节点间构成角度】
def remake_zhou_angle():
    for i1 in range(0,len(patient_list)):                                                                   #遍历各病人
        for i2 in range(0,patient_list[i1].period_length):                                                  #遍历该病人各时期
            for i3 in range(0,len(patient_list[i1].period_list[i2].stent_list)):                            #遍历病人支架
                patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list.clear()                     #重新计算前将数据清零
                if i3==0:
                    all_angle=0     
                    for i4 in range(1,len(patient_list[i1].period_list[i2].stent_list[i3].huan_list)-1):     #遍历病人支架环（去除首尾两点仅对中间角度进行运算）
                        angle_temp=round(angle(
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4-1].point_mid,
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_mid,
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4+1].point_mid,
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_mid
                        ),2)
                        all_angle=all_angle+angle_temp
                        patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list.append(angle_temp)
                    patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_all=all_angle
                else:
                    all_angle=0     
                    for i4 in range(1,len(patient_list[i1].period_list[i2].stent_list[i3].huan_list)-1):        #遍历病人支架环（去除首尾两点仅对中间角度进行运算）
                        angle_temp=round(angle(
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4-1].point_mid,
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_mid,
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4+1].point_mid,
                            patient_list[i1].period_list[i2].stent_list[i3].huan_list[i4].point_mid
                        ),2)
                        all_angle=all_angle+angle_temp
                        patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list.append(angle_temp)
                    patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_all=all_angle
        tian=HowLong(patient_list[i1].period_list[0].data,patient_list[i1].period_list[patient_list[i1].period_length-1].data)  #记录病例首尾时间
    return 

#输出支架夹角数据
def write_zhou_angle():
    remake_zhou_angle()                                                                                     #重新计算支架单体轴向夹角
    #打开 重写show_database文件
    write_workbook = xlsxwriter.Workbook(route+"\\zhou_angle.xlsx")
    write_worksheet = write_workbook.add_worksheet('Sheet1')
    number=0
    for i1 in range(0,len(patient_list)):                                                                   #遍历各病人
        write_worksheet.write(number,0,str(patient_list[i1].name))#写入病人姓名
        pianyi=0
        for i2 in range(0,patient_list[i1].period_length):                                                  #遍历该病人各时期
            write_worksheet.write(number+2,i2,patient_list[i1].period_list[i2].data)#写入病人时期
            for i3 in range(0,len(patient_list[i1].period_list[i2].stent_list)):                            #遍历病人支架
                l1=len(patient_list[i1].period_list[i2].stent_list[0].huan_list)
                pianyi=l1
                if i3==0:
                    write_worksheet.write(number+1,i2,"近端支架")#写入支架类型
                    write_worksheet.write(number+1+l1,i2,"和值")#写入和值
                    all_angle=0     
                    for i4 in range(0,len(patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list)):        #遍历病人支架环（去除首尾两点仅对中间角度进行运算）
                        angle_temp=patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list[i4]
                        all_angle=all_angle+angle_temp
                        write_worksheet.write(number+3+i4,i2,round(angle_temp,2))
                    write_worksheet.write(number+2+l1,i2,round(all_angle,2))
                else:
                    l2=len(patient_list[i1].period_list[i2].stent_list[1].huan_list)
                    pianyi=pianyi+l2
                    write_worksheet.write(number+3+l1,i2,"远端支架")#写入支架类型
                    write_worksheet.write(number+4+l1,i2,patient_list[i1].period_list[i2].data)#写入病人时期
                    write_worksheet.write(number+3+l1+l2,i2,"和值")#写入和值
                    all_angle=0     
                    for i4 in range(0,len(patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list)):        #遍历病人支架环（去除首尾两点仅对中间角度进行运算）
                        angle_temp=patient_list[i1].period_list[i2].stent_list[i3].zhou_angle_list[i4]
                        all_angle=all_angle+angle_temp
                        write_worksheet.write(number+5+l1+i4,i2,round(angle_temp,2))
                    write_worksheet.write(number+4+l1+l2,i2,round(all_angle,2))
        number=number+pianyi+6
    write_workbook.close()
    return 

#支架夹角数据占比分析   #输出支架夹角数据的统计量
def write_statistics_zhou_angle():
    #打开 重写show_statistics_zhou_angle文件
    write_workbook = xlsxwriter.Workbook(route+"\\show_statistics_zhou_angle.xlsx")
    write_worksheet = write_workbook.add_worksheet('Sheet1')
    number=0            #偏移量记录

    write_worksheet.write(number,0,"非限制型支架")
    write_worksheet.write(number,1,"日期间距")
    write_worksheet.write(number,2,"近端支架和值平均变化速率")
    write_worksheet.write(number,3,"限制型支架")
    write_worksheet.write(number,4,"日期间距")
    write_worksheet.write(number,5,"近端支架和值平均变化速率")
    write_worksheet.write(number,6,"远端支架和值平均变化速率")

    number=1
    
    number_a=0          #记录非限制型
    number_b=0          #记录限制型
    for i1 in range(0,len(patient_list)):                                                                   #遍历各病人
        if patient_list[i1].xianzhixing==False: #若为非限制性支架
            tian=HowLong(patient_list[i1].period_list[0].data,patient_list[i1].period_list[patient_list[i1].period_length-1].data)  #记录病例首尾时间
            name=patient_list[i1].name
            delta_zhou_angle=patient_list[i1].period_list[patient_list[i1].period_length-1].stent_list[0].zhou_angle_all - patient_list[i1].period_list[0].stent_list[0].zhou_angle_all
            v_change=delta_zhou_angle/tian
            write_worksheet.write(number+number_a,0,str(name))
            write_worksheet.write(number+number_a,1,tian)
            write_worksheet.write(number+number_a,2,v_change)
            number_a=number_a+1
        else:                                   #若为限制型支架
            tian=HowLong(patient_list[i1].period_list[0].data,patient_list[i1].period_list[patient_list[i1].period_length-1].data)  #记录病例首尾时间
            name=patient_list[i1].name
            delta_zhou_angle1=patient_list[i1].period_list[patient_list[i1].period_length-1].stent_list[0].zhou_angle_all - patient_list[i1].period_list[0].stent_list[0].zhou_angle_all
            delta_zhou_angle2=patient_list[i1].period_list[patient_list[i1].period_length-1].stent_list[1].zhou_angle_all - patient_list[i1].period_list[0].stent_list[1].zhou_angle_all
            v_change1=delta_zhou_angle1/tian
            v_change2=delta_zhou_angle2/tian
            write_worksheet.write(number+number_b,3,str(name))
            write_worksheet.write(number+number_b,4,tian)
            write_worksheet.write(number+number_b,5,v_change1)
            write_worksheet.write(number+number_b,6,v_change2)
            number_b=number_b+1
        

    write_workbook.close()
    return 

#支架环模拟半径分析     #输出支架首尾环形状数据 使用六边形圈图进行绘图
def write_huan_shape():
    #打开 重写huan_shape文件
    write_workbook = xlsxwriter.Workbook(route+"\\huan_shape.xlsx")
    write_worksheet = write_workbook.add_worksheet('Sheet1')
    number=0 
    number_temp=0
    for i1 in range(0,len(patient_list)):                                                                   #遍历各病人
        write_worksheet.write(number,0,str(patient_list[i1].name))#写入病人姓名
        pianyi=0
        for i2 in range(0,patient_list[i1].period_length):                                                  #遍历该病人各时期
            number_temp=number
            pianyi=0
            for i3 in range(0,len(patient_list[i1].period_list[i2].stent_list)):
                F_huan=patient_list[i1].period_list[i2].stent_list[i3].huan_list[0]     #首环
                L_huan=patient_list[i1].period_list[i2].stent_list[i3].huan_list[-1]    #尾环
                pianyi=pianyi+F_huan.number_points+L_huan.number_points+6
                #首环打印
                str_temp=ALL_stent_type[int(patient_list[i1].period_list[i2].stent_list[i3].stent_type)]+"首环"
                write_worksheet.write(number_temp+1,i2,str_temp)#写入支架环类型
                write_worksheet.write(number_temp+2,i2,patient_list[i1].period_list[i2].data)#写入病人时期
                for i4 in range(0,F_huan.number_points):
                    write_worksheet.write(number_temp+3+i4,i2,F_huan.r_list[i4])
                number_temp=number_temp+2+F_huan.number_points
                #尾环打印
                str_temp=ALL_stent_type[int(patient_list[i1].period_list[i2].stent_list[i3].stent_type)]+"尾环"
                write_worksheet.write(number_temp+1,i2,str_temp)#写入支架环类型
                write_worksheet.write(number_temp+2,i2,patient_list[i1].period_list[i2].data)#写入病人时期
                for i4 in range(0,L_huan.number_points):
                    write_worksheet.write(number_temp+3+i4,i2,L_huan.r_list[i4])
                number_temp=number_temp+2+L_huan.number_points
        number=number+pianyi        #number移位

    write_workbook.close()
    return 


#内存数据初始化部分：
read(route+"\\database.xlsx")   #读取自身database 
arrangement()                   #进行整理

#本地数据库初始化部分：
#db = pymysql.connect(host="localhost",user="root",password="baimatengqq2",database="database_mysql")                # 打开数据库链接
#cursor = db.cursor()                                                                                                # 使用 cursor() 方法创建一个游标对象 cursor

#【网络交互函数】

#本地信息展示函数
def show_all_inf():
    print("the range is ",len(inf_list))
    for i in range(0,len(inf_list)):
        print("IP:",inf_list[i].ip,"socket:",inf_list[i].sock,"ID",inf_list[i].id)
    return 

#暂存展示函数
def show_all_inf_temp():
    print("the range is ",len(inf_temp_list))
    for i in range(0,len(inf_temp_list)):
        print("IP:",inf_temp_list[i].ip,"socket:",inf_temp_list[i].sock,"ID",inf_temp_list[i].id)
    return 

#接收线程
def Sever(): 
    s = socket(AF_INET, SOCK_STREAM)
    s.bind((inf_self.ip,inf_self.sock)) 
    s.listen(5) 

    while True: 
        #信息标识码 1 : 添加LAN身份请求
        #信息标识码 2 : 组网传输请求
        #信息标识码 3 : database数据共享请求
        #print("waitting for a new connection")
        conn, addr = s.accept()                                     #等待链接 阻塞本线程
        #print("Accept new connection from %s:%s" % addr) 
        sentence=conn.recv(1024).decode()                           #信息标识码读取
        temp_list=sentence.split()
        if temp_list[0]=="1":         #信息标识码 1 : 添加LAN身份请求
            #print("添加LAN身份请求")
            inf_temp=inf(temp_list[1],temp_list[2],temp_list[3])    #创建信息实例
            inf_list.append(inf_temp)                               #追加该身份信息
            show_all_inf()
            #print("更新完成")
        elif temp_list[0]=='2':       #信息标识码 2 : 组网传输请求
            #print("组网传输请求")
            inf_temp_list.clear()   #清空原有temp_ip表
            for i in range(0,int((len(temp_list)-1)/3)):
                inf_temp=inf(temp_list[3*i+1],temp_list[3*i+2],temp_list[3*i+3])    #创建信息实例
                inf_temp_list.append(inf_temp)                                      #追加该身份信息
            show_all_inf_temp()
            #print("更新完成")
        elif temp_list[0]=='3':       #信息标识码 3 : database数据共享请求
            #print("数据共享请求")
            filesize = str(os.path.getsize(route+"\\database.xlsx"))         #获取本地database文件大小
            conn.send(filesize.encode())                                     #传输本地database文件大小
            f = open(route+"\\database.xlsx",'rb')                           #打开本地database文件
            for line in f:                                                   #传输本地database文件
                conn.send(line)
            f.close()                                                        #进行文件关闭
        elif temp_list[0]=='4':       #信息标识码 4 : 洪泛搜索请求
            patient_id=temp_list[-1]                #取出目标病人id
            org_ip=temp_list[-3]                    #取出原始主机的ip
            org_socket=int(temp_list[-2])           #取出原始主机的端口号
            have_patient_id=False                   #查询本地病人信息id标识
            TTL=int(temp_list[1])                   #取出TTL值
            inf_ed_list=[]                          #取出已访问inf列表
            for i in range(0,int((len(temp_list)-3)/2)):
                inf_temp=inf(temp_list[2+2*i],temp_list[2+2*i+1],"unknown")
                inf_ed_list.append(inf_temp)
            patient_temp=''                         #暂存病人数据               
            for i in range(0,len(patient_list)):
                if patient_list[i].ID == patient_id:        #若发现所需id
                    patient_temp=patient_list[i]            #病例赋值
                    have_patient_id=True                    #更改标识
                    break
            if have_patient_id==True:                       #尝试写入temp.xlsx并传回数据 之后删除本地temp文件
                #组建请求报文
                data0="5"                                   #信息标识码 5 : 洪泛返回链接
                data0=data0+" "+inf_self.ip+" "+str(inf_self.sock)+" "+inf_self.id+" "+patient_id      #data0组建
                target_ip=org_ip                            #获取目标ip
                target_socket=org_socket                    #获取目标端口号
                s_temp = socket(AF_INET, SOCK_STREAM)
                flag=s_temp.connect_ex((target_ip, int(target_socket)))       #发起TCP链接
                jishu=0
                name="flooding_temp.xlsx"                                #本地中间传输文件命名
                patient_temp.write(route,name)                           #向本地文件夹写入中间传输文件
                filesize = str(os.path.getsize(route+"\\"+name))         #获取本地中间传输文件大小
                while True:
                    jishu=jishu+1
                    if flag==0:
                        s_temp.send(data0.encode())                              #传输指令标识信息
                        s_temp.send(filesize.encode())                           #传输中间传输文件大小
                        f = open(route+"\\"+name,'rb')                           #打开中间传输文件
                        for line in f:                                           #传输本地database文件
                            s_temp.send(line)
                        f.close()                   #关闭文件
                        s_temp.close()              #发送端成功传输信息，关闭TCP链接
                        os.remove(route+"\\"+name)  #删除中间文件
                        break
                    else:
                        if jishu>3:
                            break
                        flag=s_temp.connect_ex((target_ip, int(target_socket)))      #重新发起TCP链接
            else:                                   #判断TTL-1情况 检索发送序列 增添自己标识后将其发送
                TTL=TTL-1
                if TTL>0:                           #TTL仍然存活，检索发送序列进行发送
                    data0="4"                           #信息标识码 4 : 洪泛请求连接
                    data0=data0+" "+str(TTL)+" "+inf_self.ip+" "+str(inf_self.id)   #data0加入自身信息
                    for i in range(2,len(temp_list)):
                        data0=data0+" "+temp_list[i]                                #data0加入原始信息
                    for i in range(0,len(inf_flooding_list)):                            #该节点遍历洪泛peer列表
                        target_ip=inf_flooding_list[i].ip                                #获取目标ip
                        target_socket=inf_flooding_list[i].sock                          #获取目标端口号
                        same_flag=False                                                  #重复性检查标志位
                        for i1 in range(0,len(inf_ed_list)):                             #遍历已经过节点列表进行重复性检测
                            if target_ip==inf_ed_list[i1].ip and target_socket==inf_ed_list[i1].sock:
                                same_flag==True                                          #该节点已经历过查找
                                break
                        if same_flag==False:                                             #若该节点从未经历过查找
                            s_temp = socket(AF_INET, SOCK_STREAM)
                            flag=s_temp.connect_ex((target_ip, int(target_socket)))      #发起TCP链接
                            jishu=0
                            while True:
                                jishu=jishu+1
                                if flag==0:
                                    s_temp.send(data0.encode())                          #传输指令标识信息
                                    s_temp.close()              #发送端成功传输信息，关闭TCP链接
                                    break
                                else:
                                    if jishu>3:
                                        break
                                    flag=s.connect_ex((target_ip, int(target_socket)))   #重新发起TCP链接
        elif temp_list[0]=='5':       #信息标识码 5 : 洪泛文件传回请求
            from_ip=temp_list[-4]                           #来源ip储存
            from_socket=temp_list[-3]                       #来源sock储存
            from_id=temp_list[-2]                           #来源id储存
            patient_id=temp_list[-1]                        #传回patient_id储存
            data_range = conn.recv(1024)                    #首次获取传输文件长度信息
            file_total_size = int(data_range.decode())      #获取传输文件长度
            received_size = 0                               #记录已获取长度
            name="flooding_temp.xlsx"                       #中间传输文件名
            f = open(route+"\\"+name, 'wb')                 #打开临时储存文件
            while received_size < file_total_size:          
                data = conn.recv(1024)
                f.write(data)
                received_size += len(data)  
            f.close()                                       #关闭文件
            read(route+"\\"+name)                           #进行文件读取 和 校验
            os.remove(route+"\\"+name)                      #删除临时文件
            print("")
            print("recieve:",patient_id,"from",from_ip,from_socket,from_id)     #回执信息打印

#分布式网络共享模块
thread_sever=threading.Thread(target=Sever, args=())
thread_sever.start()


#数据库模块

# 构建delete_db函数 实现病人信息的级联删除
def delete_db(ID):
    sql_1="delete from patient where patient.ID="+ID
    try:
        # 执行sql语句
        cursor.execute(sql_1)
        # 提交到数据库执行
        db.commit()
    except:
        print("ERROR in",sql_1)
        db.rollback

# 构建delete_db_all函数 实现数据库清空
def delete_db_all():
    sql_1="delete from patient"
    try:
        # 执行sql语句
        cursor.execute(sql_1)
        # 提交到数据库执行
        db.commit()
    except:
        print("ERROR in",sql_1)
        db.rollback

# 构建ini_db函数 使用本地内存数据 初始化数据库信息
def ini_db():
    delete_db_all()                      #级联删除数据库所有信息
    for i1 in range(0,len(patient_list)):                           #遍历病人列表
        patient_temp=patient_list[i1]                               #取出病人实例
        name=patient_temp.name                                      #姓名赋值
        sex=patient_temp.sex                                        #性别赋值
        ID=patient_temp.ID                                          #ID赋值
        calculation=patient_temp.calculation                        #可计算性赋值
        xianzhixing=patient_temp.xianzhixing                        #限制型判别赋值
        sql_1="insert into patient values(\""+ID+"\",\""+sex+"\",\""+name+"\","+str(calculation)+","+str(xianzhixing)+")"
        try:
            # 执行sql语句
            cursor.execute(sql_1)
            # 提交到数据库执行
            db.commit()
        except:
            print("ERROR in:",name,ID,sql_1)
            db.rollback
            continue
        
        for i2 in range(0,len(patient_temp.period_list)):
            period_temp=patient_temp.period_list[i2]                        #取特定时期
            period_ID=ID+"."+str(i2)                                        #构造时期ID
            data=period_temp.data                                           #取data
            top12_x=period_temp.top12_x                                     #取top12_x坐标
            top12_y=period_temp.top12_y                                     #取top12_y坐标
            top12_z=period_temp.top12_z                                     #取top12_z坐标
            sql_2_1="insert into patient_period values(\""+ID+"\",\""+period_ID+"\""+")"
            sql_2_2="insert into period values(\""+period_ID+"\",\""+data+"\","+str(top12_x)+","+str(top12_y)+","+str(top12_z)+")"
            try:
                cursor.execute(sql_2_1)
                db.commit()
            except:
                print("ERROR in:",sql_2_1)
                db.rollback
                continue
            try:
                cursor.execute(sql_2_2)
                db.commit()
            except:
                print("ERROR in:",sql_2_2)
                db.rollback
                continue
            
            for i3 in range(0,len(period_temp.stent_list)):
                stent_temp=period_temp.stent_list[i3]
                stent_ID=period_ID+"."+str(i3)
                stent_type=stent_temp.stent_type
                stent_shape=stent_temp.stent_shape
                sql_3_1="insert into period_stent values(\""+period_ID+"\",\""+stent_ID+"\""+")"
                sql_3_2="insert into stent values(\""+stent_ID+"\","+str(stent_type)+","+str(stent_shape)+")"
                try:
                    cursor.execute(sql_3_1)
                    db.commit()
                except:
                    print("ERROR in:",sql_3_1)
                    db.rollback
                    continue
                try:
                    cursor.execute(sql_3_2)
                    db.commit()
                except:
                    print("ERROR in:",sql_3_2)
                    db.rollback
                    continue
                
                for i4 in range(0,len(stent_temp.huan_list)):
                    huan_temp=stent_temp.huan_list[i4]
                    huan_ID=stent_ID+"."+str(i4)
                    flag=huan_temp.flag
                    point_mid_x=huan_temp.point_mid.x
                    point_mid_y=huan_temp.point_mid.y
                    point_mid_z=huan_temp.point_mid.z
                    sql_4_1="insert into stent_huan values(\""+stent_ID+"\",\""+huan_ID+"\""+")"
                    sql_4_2="insert into huan values(\""+huan_ID+"\",\""+flag+"\","+str(point_mid_x)+","+str(point_mid_y)+","+str(point_mid_z)+")"
                    try:
                        cursor.execute(sql_4_1)
                        db.commit()
                    except:
                        print("ERROR in:",sql_4_1)
                        db.rollback
                        continue
                    try:
                        cursor.execute(sql_4_2)
                        db.commit()
                    except:
                        print("ERROR in:",sql_4_2)
                        db.rollback
                        continue

# 提供更改任意ID的calculation标识方法
def change_patient_calculation_db(ID,calculation_temp):
    sql_1="UPDATE patient SET calculation="+str(calculation_temp)+" where ID="+ID
    try:
        cursor.execute(sql_1)
        db.commit()
    except:
        print("ERROR in",sql_1)
        db.rollback

# 提供更改任意ID的xianzhixing标识方法
def change_patient_xianzhixing_db(ID,xianzhixing_temp):
    sql_1="UPDATE patient SET xianzhixing="+str(xianzhixing_temp)+" where ID="+ID
    try:
        cursor.execute(sql_1)
        db.commit()
    except:
        print("ERROR in",sql_1)
        db.rollback

# 查询函数 提供直接查询方法
def select_SQL(str_sql):
    try:
        cursor.execute(str_sql)
        data = cursor.fetchall()
        print(data)
        db.commit()
    except:
        print("ERROR in",str_sql)
        db.rollback

#查询函数 提供病人信息查询方法
def select_patient_db(ID):
    sql_1="select * from patient where ID="+str(ID)
    try:
        cursor.execute(sql_1)
        data = cursor.fetchall()
        print(data)
        db.commit()
    except:
        print("ERROR in",sql_1)
        db.rollback






'''
# 系统登录部分
host_temp=input("please input the name:")
password_temp=input("please input the password:")
flag_temp=1
while True:
    sql_1="select * from user where name="+"\""+host_temp+"\""+" and password="+"\""+password_temp+"\""
    cursor.execute(sql_1)
    data = cursor.fetchall()
    if len(data)==0:
        print("password or name ERROR!!!")
        host_temp=input("please input the host:")
        password_temp=input("please input the password:")
    else:
        print("welcome back sir")
        jurisdiction.append(data[0][2])
        jurisdiction.append(data[0][3])
        jurisdiction.append(data[0][4])
        jurisdiction.append(data[0][5])
        jurisdiction.append(data[0][6])
        break
'''

#发送线程 [同时负责程序控制]

while True:
    # 发送端设置:
    # show_database                     显示本地database
    # show_3D                           显示本地3D文件
    # write_zhou_angle                  写出支架轴向夹角文件 zhou_angle
    # write_statistics_zhou_angle       写出支架轴向夹角文件 statistics_zhou_angle
    # write_huan_shape                  写出支架首尾环文件   huan_shape
    # -1   - 查看本地ip表 ip-temp表
    # 0    - 发送添加身份请求
    # 1    - LAN发送组网传输回馈 
    # 2.1  - database模块数据传输事务添加请求 使用ip-list
    # 2.2  - database模块数据传输事务添加请求 使用ip-temp-list
    # 3    - 洪泛搜索特定ID患者信息请求       使用inf_flooding_list

    a = input("please input oder:")
    if a == "-1":
        print("ip-list")
        show_all_inf()
        print("ip-temp-list")
        show_all_inf_temp()
    elif a == "0":        #添加身份请求 [输入目标主机IP与SOCKET进行身份信息添加]
        target_ip=input("please input target_ip:")                  #输入目标ip
        target_socket=input("please input target_socket:")          #输入目标端口号
        s = socket(AF_INET, SOCK_STREAM)
        data="1"+" "+inf_self.ip+" "+str(inf_self.sock)+" "+inf_self.id     #组建传输信息

        flag=s.connect_ex((target_ip, int(target_socket)))                          #发起TCP链接
        jishu=0
        while True:
            jishu=jishu+1
            if flag==0:
                s.send(data.encode())
                s.close()                   #发送端成功传输信息，关闭TCP链接
                jishu=0                     #jishu=0表明正常发送
                break
            else:
                if jishu>3:
                    jishu=-1                #jishu=-1表明发送3次失败，放弃本次传输
                    break
                time.sleep(1)
                flag=s.connect_ex((target_ip, int(target_socket)))                  #重新发起TCP链接
        if jishu==0:
            print(target_ip,target_socket,"successfully sent")
        elif jishu==-1:
            print(target_ip,target_socket,"fail in send")
    elif a == "1":        #LAN发送组网传输回馈 [将本地inf_list表广播至本网络节点]
        #组建data报文
        data='2'
        for i in range(0,len(inf_list)):
            data=data+" "+inf_list[i].ip+" "+str(inf_list[i].sock)+" "+inf_list[i].id
        #进行TCP文件传输
        for i in range(0,len(inf_list)):
            if inf_list[i].ip==inf_self.ip and inf_list[i].sock==inf_self.sock:
                continue
            else:
                target_ip=inf_list[i].ip
                target_socket=inf_list[i].sock
                s = socket(AF_INET, SOCK_STREAM)

                flag=s.connect_ex((target_ip, int(target_socket)))                          #发起TCP链接
                jishu=0
                while True:
                    jishu=jishu+1
                    if flag==0:
                        s.send(data.encode())
                        s.close()                   #发送端成功传输信息，关闭TCP链接
                        jishu=0                     #jishu=0表明正常发送
                        break
                    else:
                        if jishu>3:
                            jishu=-1                #jishu=-1表明发送3次失败，放弃本次传输
                            break
                        time.sleep(0.1)
                        flag=s.connect_ex((target_ip, int(target_socket)))                  #重新发起TCP链接
                if jishu==0:
                    print(target_ip,target_socket,"successfully sent")
                elif jishu==-1:
                    print(target_ip,target_socket,"fail in send")
    elif a == '2.1':      #请求进行database共享 按ip-list寻址
        #组建请求报文
        data0='3' 
        str0=" temp temp"
        data0=data0+" temp temp"
        data0=str(data0)
        #进行TCP文件传输
        for i in range(0,len(inf_list)):
            if inf_list[i].ip==inf_self.ip and inf_list[i].sock==inf_self.sock:
                continue
            else:
                target_ip=inf_list[i].ip
                target_socket=inf_list[i].sock
                s = socket(AF_INET, SOCK_STREAM)
                flag=s.connect_ex((target_ip, int(target_socket)))                          #发起TCP链接
                jishu=0
                while True:
                    jishu=jishu+1
                    if flag==0:                         #连接正常 开始进行database传输 获得的database存放于database-temp中 用完即删
                        s.send(data0.encode())          #发送标识信息 信息标识码 - 3
                        data_range = s.recv(1024)       #首次获取传输文件长度信息
                        file_total_size = int(data_range.decode())       #获取传输文件长度
                        received_size = 0                                #记录已获取长度
                        f = open(route+"\\database_temp.xlsx", 'wb')              #打开临时储存文件
                        while received_size < file_total_size:           #未接收完时：
                            data = s.recv(1024)
                            f.write(data)
                            received_size += len(data)
                        #接收完成 关闭TCP 与文件
                        s.close()
                        f.close()
                        jishu=0
                        #进行文件读取 
                        read(route+"\\database_temp.xlsx")
                        #进行文件校验
                        arrangement()
                        #进删除临时文件
                        os.remove(route+"\\database_temp.xlsx")
                        break
                    else:
                        if jishu>3:
                            jishu=-1
                            break
                        time.sleep(0.1)
                        flag=s.connect_ex((target_ip, int(target_socket)))                   #重新发起TCP链接
                if jishu==0:
                    print(target_ip,target_socket,"successfully sent")
                elif jishu==-1:
                    print(target_ip,target_socket,"fail in send")
    elif a == '2.2':      #请求进行database共享 按ip-temp-list寻址
        #组建请求报文
        data0='3' 
        str0=" temp temp"
        data0=data0+" temp temp"
        data=str(data0)
        #进行TCP文件传输
        for i in range(0,len(inf_temp_list)):
            if inf_temp_list[i].ip==inf_self.ip and inf_temp_list[i].sock==inf_self.sock:
                continue
            else:
                target_ip=inf_temp_list[i].ip
                target_socket=inf_temp_list[i].sock
                s = socket(AF_INET, SOCK_STREAM)
                flag=s.connect_ex((target_ip, int(target_socket)))                          #发起TCP链接
                jishu=0
                while True:
                    jishu=jishu+1
                    if flag==0:                         #连接正常 开始进行database传输 获得的database存放于database-temp中 用完即删
                        s.send(data0.encode())          #发送标识信息 信息标识码 - 3
                        data_range = s.recv(1024)       #首次获取传输文件长度信息
                        file_total_size = int(data_range.decode())       #获取传输文件长度
                        received_size = 0                                #记录已获取长度
                        f = open(route+"\\database_temp.xlsx", 'wb')              #打开临时储存文件
                        while received_size < file_total_size:           #未接收完时：
                            data = s.recv(1024)
                            f.write(data)
                            received_size += len(data)
                        #接收完成 关闭TCP 关闭文件
                        s.close()
                        f.close()
                        jishu=0
                        #进行文件读取 和 校验
                        read(route+"\\database_temp.xlsx")
                        #删除临时文件
                        os.remove(route+"\\database_temp.xlsx")
                        break
                    else:
                        if jishu>3:
                            jishu=-1
                            break
                        time.sleep(0.1)
                        flag=s.connect_ex((target_ip, int(target_socket)))                   #重新发起TCP链接
                if jishu==0:
                    print(target_ip,target_socket,"successfully sent")
                elif jishu==-1:
                    print(target_ip,target_socket,"fail in send")
    elif a == "3":        #洪泛网络搜寻特定ID数据
        #组建请求报文
        data0="4"                                               #信息标识码4 ：洪泛数据请求
        TTL_temp=input("please define the TTL:")              #定义TTL
        patient_id=input("please input the patient's ID:")    #输入病人ID
        data0=data0+" "+str(TTL_temp)+" "+inf_self.ip+" "+str(inf_self.sock)+" "+patient_id
        for i in range(0,len(inf_flooding_list)):
            target_ip=inf_flooding_list[i].ip                   #获取目标ip
            target_socket=inf_flooding_list[i].sock             #获取目标端口号
            s = socket(AF_INET, SOCK_STREAM)
            flag=s.connect_ex((target_ip, int(target_socket)))  #发起TCP链接
            jishu=0
            while True:
                jishu=jishu+1
                if flag==0:
                    s.send(data0.encode())
                    s.close()                   #发送端成功传输信息，关闭TCP链接
                    jishu=0                     #jishu=0表明正常发送
                    break
                else:
                    if jishu>3:
                        jishu=-1                #jishu=-1表明发送3次失败，放弃本次传输
                        break
                    time.sleep(1)
                    flag=s.connect_ex((target_ip, int(target_socket)))                  #重新发起TCP链接
            if jishu==0:
                print(target_ip,target_socket,"successfully sent")
            elif jishu==-1:
                print(target_ip,target_socket,"fail in send")
    elif a == "show_database":
        show_database()
    elif a == "show_3D":
        show_3D()
    elif a == "write_zhou_angle":
        write_zhou_angle()
    elif a == "write_statistics_zhou_angle":
        write_statistics_zhou_angle()
    elif a == "write_huan_shape":
        write_huan_shape()
    elif a == "show_flooding_list":
        for i in range(0,len(inf_flooding_list)):
            print("IP:",inf_flooding_list[i].ip,"socket:",inf_flooding_list[i].sock,"ID",inf_flooding_list[i].id)
    elif a == "clear_flooding_list":
        inf_flooding_list.clear()
    elif a == "append_flooding_list":
        peer_ip=input("please input peer_ip:")                  #输入目标ip
        peer_socket=input("please input peer_socket:")          #输入目标端口号
        inf_temp=inf(peer_ip,int(peer_socket),"unknown")
        inf_flooding_list.append(inf_temp)                      #追加peer_temp
    elif a == "ini_db":
        #权限检测
        if jurisdiction[0]==1 and jurisdiction[1]==1 and jurisdiction[2]==1 and jurisdiction[3]==1:
            ini_db()
        else:
            print("no permission to access")
    elif a == "show_permission_db":
        print("insert",jurisdiction[0])
        print("delete",jurisdiction[1])
        print("update",jurisdiction[2])
        print("drop",jurisdiction[3])
        print("select",jurisdiction[4])
    elif a == "change_patient_calculation_db":
        #权限检测
        if jurisdiction[2]==1 and jurisdiction[4]==1:
            ID_temp=input("please input patient's ID:")
            calculation_temp=input("please input calculation flag:")
            change_patient_calculation_db(ID_temp,calculation_temp)
    elif a == "change_patient_xianzhixing_db":
        if jurisdiction[2]==1 and jurisdiction[4]==1:
            ID_temp=input("please input patient's ID:")
            xianzhixing_temp=input("please input xianzhixing flag:")
            change_patient_calculation_db(ID_temp,xianzhixing_temp)
    elif a == "select_SQL_db":
        if jurisdiction[0]==1 and jurisdiction[1]==1 and jurisdiction[2]==1 and jurisdiction[3]==1 and jurisdiction[4]==1:
            sql_1=input("please input SQL:")
            select_SQL(sql_1)
    elif a == "select_patient_db":
        if jurisdiction[4]==1:
            ID_temp=input("please input patient's ID:")
            select_patient_db(ID_temp)
    elif a == "show_self_roles_db":
        if jurisdiction[0]==1 and jurisdiction[1]==1 and jurisdiction[2]==1 and jurisdiction[3]==1 and jurisdiction[4]==1:
            print("PERMISSON--PIONEER")
        elif jurisdiction[2]==1 and jurisdiction[4]==1:
            print("PERMISSON--EXPLORER")
        elif jurisdiction[4]==1:
            print("PERMISSON--USER")
    elif a == "use_func_db":
        if jurisdiction[4]==1:
            sql_1="select getnumber()"
            try:
                print("number of the patients is:",len(patient_list))
                cursor.execute(sql_1)
                db.commit()
            except:
                db.rollback

#【数据库函数】








    




#ini_db()                    # 初始化db数据库 
#delete_db("6440338")        # 提供指定删除方法
#change_patient_calculation_db("3892150",1)     # 提供更改任意ID的calculation标识方法
#change_patient_xianzhixing_db("3892150",1)     # 提供更改任意ID的xianzhixing标识方法

#sql_0=input("please input your SQL:")
#select_SQL(sql_0)                              # 提供SQL通用接入入口

#select_patient_db(35595496)                    # 提供单个查询病例方法





db.close()              # 关闭数据库连接






























